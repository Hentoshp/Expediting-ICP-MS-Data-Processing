import pandas as pd
import sys
from openpyxl import load_workbook

path = "/workspaces/131570367/project/"
def main():
    file_name_raw = input("Raw data file name: ", )
    sample_data = get_sample_data(get_raw_data(file_name_raw))
    file_name_data = input("Inital excel file name: ", )
    requested_metals = get_requested_metals(file_name_data)
    move_data(sample_data, requested_metals, file_name_data)

def get_raw_data(file_name):
    # reading in raw data excel sheet
    try:
        raw_data = pd.read_excel(path + file_name)
        return raw_data.drop(["M2", "M3", "M4", "vial position"], axis = 1)
    except FileNotFoundError:
       print("Incorrect raw data file path")
       raise FileNotFoundError


def get_requested_metals(excel_page):
        # creating dictionary for client requested metals
        try:
            page = load_workbook(excel_page)
            page_sheet_1 = page['Main Data']
            requested_metals = {}
            for row in page_sheet_1.iter_rows(values_only = True, max_col = 2, min_row = 2):
                sample, metals = row
                requested_metals[sample] = metals
            return requested_metals
        except FileNotFoundError:
            print("Incorrect data file path")
            raise FileNotFoundError

def get_sample_data(df):
    # filter raw data for sample data only
    df_filtered = df[df['sample name'].str.contains('sample|spike|page')]
    return df_filtered.drop(["Cd [114]", "Ni [62]", "Pb [206]", "Pb [207]", "Cr [53]", "Cr [60]"], axis = 1)


def check_metals(sample_key, dict, col_value):
    allowed_metals = dict[sample_key].split(",")
    return any(metals.strip() in col_value for metals in allowed_metals)


def move_data(df, dict, excel_page):
    try:
        pg1 = load_workbook(excel_page)
    except FileNotFoundError:
            raise FileNotFoundError
    data = pg1['Main Data']
    for i, (index, row) in enumerate(df.iterrows()):
        sample_key = str(row[df.columns[0]])
        for col_num, col_value in enumerate(df.columns[1:], start = 3):
            cell = data.cell(row = i + 2, column = col_num)
            value = row[col_value]
            if check_metals(sample_key, dict, col_value):
                cell.value = value
            else:
                cell.value = 'NA'
    pg1.save('excel page 1 result.xlsx')













if __name__ == "__main__":
    main()




#Required libraries Pandas and Openxyl
